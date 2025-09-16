from django.shortcuts import render
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, View
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
import csv
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from attendance.models import AttendanceRecord, SessionAttendance
from events.models import Event, EventSession
from attendees.models import Attendee


class ReportsView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Date range filters
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        event_id = self.request.GET.get('event')
        
        # Base querysets for both regular and session attendance
        regular_queryset = AttendanceRecord.objects.select_related('event', 'attendee')
        session_queryset = SessionAttendance.objects.select_related('event_session__event', 'attendee')
        
        # Apply filters to regular attendance
        if date_from:
            regular_queryset = regular_queryset.filter(timestamp__date__gte=date_from)
            session_queryset = session_queryset.filter(timestamp__date__gte=date_from)
        if date_to:
            regular_queryset = regular_queryset.filter(timestamp__date__lte=date_to)
            session_queryset = session_queryset.filter(timestamp__date__lte=date_to)
        if event_id:
            regular_queryset = regular_queryset.filter(event_id=event_id)
            session_queryset = session_queryset.filter(event_session__event_id=event_id)
        
        # Combine results
        regular_records = list(regular_queryset.order_by('-timestamp')[:50])
        session_records = list(session_queryset.order_by('-timestamp')[:50])
        
        # Combine and sort by timestamp
        all_records = regular_records + session_records
        all_records.sort(key=lambda x: x.timestamp, reverse=True)
        
        context['attendance_records'] = all_records[:100]
        context['events'] = Event.objects.filter(is_active=True).order_by('-date')
        context['total_records'] = regular_queryset.count() + session_queryset.count()
        
        # Filters for form
        context['date_from'] = date_from
        context['date_to'] = date_to
        context['selected_event'] = event_id
        
        return context


class ExportView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/export.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['events'] = Event.objects.filter(is_active=True).order_by('-date')
        return context


class ExportExcelView(LoginRequiredMixin, View):
    def get(self, request):
        # Get filters
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        event_id = request.GET.get('event')
        
        # Build querysets for both regular and session attendance
        regular_queryset = AttendanceRecord.objects.select_related('event', 'attendee')
        session_queryset = SessionAttendance.objects.select_related('event_session__event', 'attendee', 'event_session')
        
        if date_from:
            regular_queryset = regular_queryset.filter(timestamp__date__gte=date_from)
            session_queryset = session_queryset.filter(timestamp__date__gte=date_from)
        if date_to:
            regular_queryset = regular_queryset.filter(timestamp__date__lte=date_to)
            session_queryset = session_queryset.filter(timestamp__date__lte=date_to)
        if event_id:
            regular_queryset = regular_queryset.filter(event_id=event_id)
            session_queryset = session_queryset.filter(event_session__event_id=event_id)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Headers
        headers = [
            'Attendee ID', 'First Name', 'Last Name', 'Email', 'Phone',
            'Event Name', 'Session Info', 'Event Date', 'Event Location', 'Attendance Time', 'IP Address'
        ]
        
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws[f'{col_letter}1'].font = ws[f'{col_letter}1'].font.copy(bold=True)
        
        # Combine and sort data
        regular_records = list(regular_queryset.order_by('-timestamp'))
        session_records = list(session_queryset.order_by('-timestamp'))
        
        all_records = []
        
        # Add regular attendance records
        for record in regular_records:
            all_records.append({
                'attendee_id': record.attendee.attendee_id,
                'first_name': record.attendee.first_name,
                'last_name': record.attendee.last_name,
                'email': record.attendee.email,
                'phone': record.attendee.phone,
                'event_name': record.event.name,
                'session_info': 'Single Event',
                'event_date': record.event.date.strftime('%Y-%m-%d'),
                'location': record.event.location,
                'timestamp': record.timestamp,
                'ip_address': record.ip_address
            })
        
        # Add session attendance records
        for record in session_records:
            all_records.append({
                'attendee_id': record.attendee.attendee_id,
                'first_name': record.attendee.first_name,
                'last_name': record.attendee.last_name,
                'email': record.attendee.email,
                'phone': record.attendee.phone,
                'event_name': record.event_session.event.name,
                'session_info': f'Session {record.event_session.session_number} - {record.event_session.session_date}',
                'event_date': record.event_session.session_date.strftime('%Y-%m-%d'),
                'location': record.event_session.location,
                'timestamp': record.timestamp,
                'ip_address': record.ip_address
            })
        
        # Sort by timestamp
        all_records.sort(key=lambda x: x['timestamp'], reverse=True)
        
        # Data rows
        for row_num, record in enumerate(all_records, 2):
            ws[f'A{row_num}'] = record['attendee_id']
            ws[f'B{row_num}'] = record['first_name']
            ws[f'C{row_num}'] = record['last_name']
            ws[f'D{row_num}'] = record['email']
            ws[f'E{row_num}'] = record['phone']
            ws[f'F{row_num}'] = record['event_name']
            ws[f'G{row_num}'] = record['session_info']
            ws[f'H{row_num}'] = record['event_date']
            ws[f'I{row_num}'] = record['location']
            ws[f'J{row_num}'] = record['timestamp'].strftime('%Y-%m-%d %H:%M:%S')
            ws[f'K{row_num}'] = record['ip_address']
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'attendance_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response


class ExportCSVView(LoginRequiredMixin, View):
    def get(self, request):
        # Get filters
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        event_id = request.GET.get('event')
        
        # Build queryset
        queryset = AttendanceRecord.objects.select_related('event', 'attendee')
        
        if date_from:
            queryset = queryset.filter(timestamp__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(timestamp__date__lte=date_to)
        if event_id:
            queryset = queryset.filter(event_id=event_id)
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        filename = f'attendance_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        writer = csv.writer(response)
        
        # Headers
        writer.writerow([
            'Attendee ID', 'First Name', 'Last Name', 'Email', 'Phone',
            'Event Name', 'Event Date', 'Event Location', 'Attendance Time', 'IP Address'
        ])
        
        # Data rows
        for record in queryset.order_by('-timestamp'):
            writer.writerow([
                record.attendee.attendee_id,
                record.attendee.first_name,
                record.attendee.last_name,
                record.attendee.email,
                record.attendee.phone,
                record.event.name,
                record.event.date.strftime('%Y-%m-%d'),
                record.event.location,
                record.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                record.ip_address
            ])
        
        return response